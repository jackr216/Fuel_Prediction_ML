"""
Fuel Prediction Pipeline
========================
Predicts April 2026 – March 2027 fuel spend at Company, Branch, and Vehicle level.

Approach:
  Forecast Spend = Predicted Volume (LightGBM) × Projected Price (BEIS trend)

Sections:
  1. Data Loading & Cleaning
  2. Price Projection  (BEIS weekly pump prices → 12-month forward projection)
  3. Volume Prediction (LightGBM panel model, recursive 12-step forecast)
  4. Report Output    (Excel: Company / Branch / Vehicle sheets)

Run:
  python run_pipeline.py
"""

import sys
sys.stdout.reconfigure(encoding='utf-8')

import warnings
warnings.filterwarnings('ignore')

import re
import calendar
from io import StringIO
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
import requests
from bs4 import BeautifulSoup
import lightgbm as lgb
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ────────────────────────────────────────────────────────────────────
BASE_DIR      = Path(__file__).resolve().parent
RAW_DIR       = BASE_DIR / 'Raw_Data'
PROCESSED_DIR = BASE_DIR / 'data' / 'processed'
REPORTS_DIR   = BASE_DIR / 'outputs' / 'reports'
PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

# ── Constants ─────────────────────────────────────────────────────────────────
DIESEL_TYPES   = {'Diesel', 'V Power Diesel'}
UNLEADED_TYPES = {'Unleaded - Medium octane', 'Unleaded High Perf'}
ALL_FUEL_TYPES = DIESEL_TYPES | UNLEADED_TYPES

# March 2026 actual Shell unit price per litre (anchor for projection)
PRICE_ANCHOR = {'Diesel': 1.2116, 'Unleaded': 1.0689}

# Price projection shape:
#   Months 1–PRICE_TREND_MONTHS : full BEIS trend (current market pressure)
#   Months beyond               : PRICE_STEADY_MONTHLY_CHG (market settles)
PRICE_TREND_MONTHS       = 3      # months of aggressive movement
PRICE_STEADY_MONTHLY_CHG = 0.002  # ~0.2%/month once prices stabilise

# EV fleet electrification adjustment:
#   The company targets ~15%/year volume reduction via EV transition.
#   Conservative estimate applied to forecasts = 10% annual volume reduction.
#   Applied as a compounding monthly factor on top of the LightGBM predictions.
#   Increase EV_ANNUAL_VOLUME_REDUCTION toward 0.15 as EV adoption accelerates.
EV_ANNUAL_VOLUME_REDUCTION = 0.10   # 10% per year = ~0.87% per month compound

# Forecast window
FORECAST_MONTHS = pd.period_range('2026-04', periods=12, freq='M')

# Minimum months of history for a vehicle to be included in forecasts
MIN_HISTORY_MONTHS = 3

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 1 — DATA LOADING & CLEANING
# ─────────────────────────────────────────────────────────────────────────────

def section1_load_data():
    print("\n" + "=" * 60)
    print("SECTION 1 — Data Loading & Cleaning")
    print("=" * 60)

    # ── Transactions ─────────────────────────────────────────────
    print("Loading fuel transactions...")
    txn = pd.read_excel(RAW_DIR / 'Fuel_Transactions.xlsx')
    print(f"  {len(txn):,} rows loaded")

    # Filter to relevant fuel types
    txn = txn[txn['ProductName'].isin(ALL_FUEL_TYPES)].copy()
    print(f"  {len(txn):,} rows after fuel-type filter")

    # Standardise fuel type
    txn['fuel_type'] = txn['ProductName'].apply(
        lambda x: 'Diesel' if x in DIESEL_TYPES else 'Unleaded'
    )

    # Parse dates → year_month Period
    txn['TransactionDate'] = pd.to_datetime(txn['TransactionDate'], errors='coerce')
    txn['year_month'] = txn['TransactionDate'].dt.to_period('M')

    # Unit price per litre (handle divide-by-zero)
    txn['unit_price'] = (
        txn['Net Amount'] / txn['Litres'].replace(0, np.nan)
    ).replace([np.inf, -np.inf], np.nan)

    # Flag vehicle vs card-only rows
    txn['reg'] = txn['RegistrationNumber'].astype(str).str.strip()
    txn['is_vehicle'] = txn['reg'] != '-'

    # ── Vehicle info ──────────────────────────────────────────────
    print("Loading vehicle info...")
    veh = pd.read_excel(RAW_DIR / 'Vehicle_Info.xlsx')
    veh = veh.rename(columns={
        'Branch Name':        'branch',
        'RegistrationNumber': 'reg',
        'Asset Type':         'asset_type',
        'Vehicle Status':     'status',
        'Registered Date':    'registered_date',
        'Make':               'make',
        'Model':              'model',
        'Derivative':         'derivative',
    })
    veh['reg'] = veh['reg'].astype(str).str.strip()
    # Keep one row per reg (latest if duplicated)
    veh = veh.drop_duplicates(subset='reg', keep='last')

    # ── Monthly aggregation — vehicle level ───────────────────────
    veh_txn = txn[txn['is_vehicle']].copy()

    monthly_veh = (
        veh_txn.groupby(['year_month', 'reg', 'Branch Name', 'fuel_type'])
        .agg(
            litres          = ('Litres',     'sum'),
            net_amount      = ('Net Amount', 'sum'),
            avg_unit_price  = ('unit_price', 'mean'),
        )
        .reset_index()
        .rename(columns={'Branch Name': 'branch'})
    )
    monthly_veh = monthly_veh.merge(
        veh[['reg', 'asset_type', 'make', 'model', 'derivative']],
        on='reg', how='left'
    )

    # ── Monthly aggregation — branch level (incl. card-only) ──────
    monthly_branch = (
        txn.groupby(['year_month', 'Branch Name', 'fuel_type'])
        .agg(
            litres          = ('Litres',     'sum'),
            net_amount      = ('Net Amount', 'sum'),
            avg_unit_price  = ('unit_price', 'mean'),
        )
        .reset_index()
        .rename(columns={'Branch Name': 'branch'})
    )

    # ── Monthly aggregation — company level ───────────────────────
    monthly_company = (
        txn.groupby(['year_month', 'fuel_type'])
        .agg(
            litres          = ('Litres',     'sum'),
            net_amount      = ('Net Amount', 'sum'),
            avg_unit_price  = ('unit_price', 'mean'),
        )
        .reset_index()
    )

    # ── Save ──────────────────────────────────────────────────────
    monthly_veh.to_csv(PROCESSED_DIR / 'monthly_vehicle.csv', index=False)
    monthly_branch.to_csv(PROCESSED_DIR / 'monthly_branch.csv', index=False)
    monthly_company.to_csv(PROCESSED_DIR / 'monthly_company.csv', index=False)

    print(f"  Vehicle rows:  {len(monthly_veh):,}")
    print(f"  Branch rows:   {len(monthly_branch):,}")
    print(f"  Company rows:  {len(monthly_company):,}")
    print("Section 1 complete.\n")

    return monthly_veh, monthly_branch, monthly_company


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 2 — PRICE PROJECTION
# ─────────────────────────────────────────────────────────────────────────────

def _fetch_beis_monthly_prices():
    """
    Scrape BEIS/DESNZ weekly road fuel prices page, download the CSV,
    and return a DataFrame with columns: year_month (Period), diesel_ppl, unleaded_ppl.
    Returns None on any failure.
    """
    page_url = 'https://www.gov.uk/government/statistical-data-sets/weekly-road-fuel-prices'
    headers  = {'User-Agent': 'Mozilla/5.0 (compatible)'}

    try:
        resp = requests.get(page_url, headers=headers, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, 'html.parser')

        # Find CSV download link
        csv_url = None
        for a in soup.find_all('a', href=True):
            href = a['href']
            if href.lower().endswith('.csv') and ('weekly' in href.lower() or 'CSV' in href):
                csv_url = href
                break
        # Fallback: any CSV attachment
        if not csv_url:
            for a in soup.find_all('a', href=True):
                href = a['href']
                if href.lower().endswith('.csv'):
                    csv_url = href
                    break

        if not csv_url:
            print("  Could not find BEIS CSV link on page.")
            return None

        if not csv_url.startswith('http'):
            csv_url = 'https://www.gov.uk' + csv_url

        print(f"  BEIS CSV: {csv_url}")
        data_resp = requests.get(csv_url, headers=headers, timeout=30)
        data_resp.raise_for_status()
        raw = data_resp.text

    except Exception as e:
        print(f"  BEIS fetch failed: {e}")
        return None

    try:
        lines = raw.splitlines()

        # Locate the header row — look for a line containing 'Date' and 'nce' (pence)
        header_idx = None
        for i, line in enumerate(lines):
            if re.search(r'\bDate\b', line, re.I) and re.search(r'pence|ppl|petrol|diesel', line, re.I):
                header_idx = i
                break
        # Fallback: first line with a 4-digit year looking like a date
        if header_idx is None:
            for i, line in enumerate(lines):
                if re.search(r'\d{2}/\d{2}/\d{4}', line):
                    header_idx = max(0, i - 1)
                    break

        if header_idx is None:
            print("  Could not locate BEIS CSV header row.")
            return None

        df = pd.read_csv(StringIO('\n'.join(lines[header_idx:])),
                         on_bad_lines='skip',
                         skipinitialspace=True)
        df.columns = [str(c).strip() for c in df.columns]

        # Identify date column
        date_col = next((c for c in df.columns if re.search(r'date', c, re.I)), None)
        if date_col is None:
            date_col = df.columns[0]

        df[date_col] = pd.to_datetime(df[date_col], dayfirst=True, errors='coerce')
        df = df.dropna(subset=[date_col])
        df['year_month'] = df[date_col].dt.to_period('M')

        # Identify diesel and unleaded columns (pump price, ppl)
        diesel_col = next(
            (c for c in df.columns if re.search(r'diesel', c, re.I)
             and re.search(r'pump|ppl|price', c, re.I)), None
        )
        unl_col = next(
            (c for c in df.columns if re.search(r'unleaded|petrol|gasoline', c, re.I)
             and re.search(r'pump|ppl|price', c, re.I)), None
        )

        if diesel_col is None or unl_col is None:
            # Looser search
            price_cols = [c for c in df.columns
                          if re.search(r'pump|ppl', c, re.I) and c != date_col]
            if len(price_cols) >= 2:
                unl_col    = price_cols[0]
                diesel_col = price_cols[1]

        if diesel_col is None or unl_col is None:
            print(f"  Cannot identify price columns. Available: {list(df.columns)}")
            return None

        df[diesel_col] = pd.to_numeric(df[diesel_col], errors='coerce')
        df[unl_col]    = pd.to_numeric(df[unl_col],    errors='coerce')

        monthly = (
            df.groupby('year_month')
            .agg(diesel_ppl=(diesel_col, 'mean'),
                 unleaded_ppl=(unl_col,  'mean'))
            .reset_index()
        )
        monthly = monthly.sort_values('year_month')
        print(f"  BEIS data: {len(monthly)} months ({monthly['year_month'].min()} – {monthly['year_month'].max()})")
        return monthly

    except Exception as e:
        print(f"  BEIS parse failed: {e}")
        return None


def section2_price_projection():
    print("=" * 60)
    print("SECTION 2 — Price Projection")
    print("=" * 60)

    beis = _fetch_beis_monthly_prices()

    # Calculate 3-month average monthly % change from BEIS data
    diesel_monthly_pct   = None
    unleaded_monthly_pct = None

    if beis is not None and len(beis) >= 4:
        # Use last 3 months available before or including March 2026
        cutoff = pd.Period('2026-03', freq='M')
        recent = beis[beis['year_month'] <= cutoff].tail(4)   # need 4 rows for 3 changes

        if len(recent) >= 4:
            d_chg = recent['diesel_ppl'].pct_change().dropna().tail(3)
            u_chg = recent['unleaded_ppl'].pct_change().dropna().tail(3)
            diesel_monthly_pct   = float(d_chg.mean())
            unleaded_monthly_pct = float(u_chg.mean())
            print(f"  Diesel   3-month avg monthly change: {diesel_monthly_pct:+.2%}")
            print(f"  Unleaded 3-month avg monthly change: {unleaded_monthly_pct:+.2%}")

    # Fallback: slight upward trend consistent with early 2026 market
    if diesel_monthly_pct is None:
        diesel_monthly_pct   = 0.0025   # ~0.25% per month
        unleaded_monthly_pct = 0.0020
        print(f"  BEIS data unavailable — using default: Diesel {diesel_monthly_pct:+.2%}/mo, Unleaded {unleaded_monthly_pct:+.2%}/mo")

    # Build price projections using two-phase curve:
    #   Phase 1 (months 1–PRICE_TREND_MONTHS): full BEIS trend (current market pressure)
    #   Phase 2 (remaining months): gentle steady rate once prices stabilise
    print(f"\n  Price curve: BEIS trend for {PRICE_TREND_MONTHS} months, then {PRICE_STEADY_MONTHLY_CHG:+.2%}/mo steady")
    diesel_price   = PRICE_ANCHOR['Diesel']
    unleaded_price = PRICE_ANCHOR['Unleaded']
    rows = []
    for i, month in enumerate(FORECAST_MONTHS, 1):
        d_chg = diesel_monthly_pct   if i <= PRICE_TREND_MONTHS else PRICE_STEADY_MONTHLY_CHG
        u_chg = unleaded_monthly_pct if i <= PRICE_TREND_MONTHS else PRICE_STEADY_MONTHLY_CHG
        diesel_price   = round(diesel_price   * (1 + d_chg), 4)
        unleaded_price = round(unleaded_price * (1 + u_chg), 4)
        rows.append({
            'year_month':     str(month),
            'diesel_price':   diesel_price,
            'unleaded_price': unleaded_price,
        })

    price_proj = pd.DataFrame(rows)
    price_proj.to_csv(PROCESSED_DIR / 'price_projection.csv', index=False)

    print("\n  Projected prices (£/litre):")
    print(f"  {'Month':<15} {'Diesel':>10} {'Unleaded':>10}")
    for _, r in price_proj.iterrows():
        print(f"  {r['year_month']:<15} {r['diesel_price']:>10.4f} {r['unleaded_price']:>10.4f}")

    print("\nSection 2 complete.\n")
    return price_proj


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 3 — VOLUME PREDICTION
# ─────────────────────────────────────────────────────────────────────────────

def _build_features(series: pd.Series) -> pd.DataFrame:
    """
    Given a time-indexed Series of monthly litres (sorted oldest→newest),
    return a DataFrame of features for the next prediction step.
    """
    n = len(series)
    values = series.values.astype(float)

    def lag(k):
        return values[-k] if n >= k else np.nan

    def roll(k):
        if n >= k:
            return float(np.mean(values[-k:]))
        elif n > 0:
            return float(np.mean(values))
        return np.nan

    # Seasonality: use the *next* month's sin/cos
    next_month = (series.index[-1] + 1).month if n > 0 else 1
    month_sin  = np.sin(2 * np.pi * next_month / 12)
    month_cos  = np.cos(2 * np.pi * next_month / 12)

    return {
        'lag_1':       lag(1),
        'lag_2':       lag(2),
        'lag_3':       lag(3),
        'lag_6':       lag(6),
        'roll_3':      roll(3),
        'roll_6':      roll(6),
        'month_sin':   month_sin,
        'month_cos':   month_cos,
    }


def _train_lgbm(train_df: pd.DataFrame, fuel: str) -> lgb.Booster:
    """Train a LightGBM model for a given fuel type."""
    sub = train_df[train_df['fuel_type'] == fuel].copy()

    # Build training rows: for each vehicle+month, create lag/roll features
    records = []
    for reg, grp in sub.groupby('reg'):
        grp = grp.sort_values('year_month')
        grp['year_month_period'] = grp['year_month'].apply(lambda x: pd.Period(x, freq='M'))
        grp = grp.set_index('year_month_period')

        for i in range(MIN_HISTORY_MONTHS, len(grp)):
            history = grp.iloc[:i]['litres']
            target  = grp.iloc[i]['litres']
            feats   = _build_features(history)
            feats['target'] = target
            records.append(feats)

    if not records:
        return None

    feat_df  = pd.DataFrame(records).dropna()
    features = ['lag_1', 'lag_2', 'lag_3', 'lag_6', 'roll_3', 'roll_6', 'month_sin', 'month_cos']
    X = feat_df[features]
    y = feat_df['target']

    model = lgb.train(
        params={
            'objective':        'regression',
            'metric':           'rmse',
            'n_estimators':     300,
            'learning_rate':    0.05,
            'num_leaves':       31,
            'min_child_samples': 10,
            'verbose':          -1,
        },
        train_set=lgb.Dataset(X, y),
        num_boost_round=300,
    )
    return model


def _forecast_vehicle(history: pd.Series, model: lgb.Booster, steps: int = 12) -> list:
    """
    Recursively forecast `steps` months ahead using the trained model.
    history: pd.Series indexed by Period, sorted oldest→newest
    """
    preds = []
    working = history.copy()

    for _ in range(steps):
        feats = _build_features(working)
        feat_df = pd.DataFrame([feats])
        features = ['lag_1', 'lag_2', 'lag_3', 'lag_6', 'roll_3', 'roll_6', 'month_sin', 'month_cos']
        pred = max(float(model.predict(feat_df[features])[0]), 0.0)
        preds.append(pred)
        # Append prediction to working history for next step
        next_period = working.index[-1] + 1
        working[next_period] = pred

    return preds


def section3_volume_prediction(monthly_veh: pd.DataFrame) -> pd.DataFrame:
    print("=" * 60)
    print("SECTION 3 — Volume Prediction (LightGBM)")
    print("=" * 60)

    # Convert year_month to Period for sorting
    monthly_veh['year_month'] = monthly_veh['year_month'].apply(
        lambda x: pd.Period(str(x), freq='M')
    )

    # Filter vehicles with MIN_HISTORY_MONTHS or more
    history_counts = monthly_veh.groupby(['reg', 'fuel_type'])['year_month'].count()
    eligible = history_counts[history_counts >= MIN_HISTORY_MONTHS].reset_index()
    eligible = eligible.rename(columns={'year_month': 'month_count'})

    print(f"  Eligible vehicle×fuel pairs: {len(eligible):,}")

    all_forecasts = []

    for fuel in ['Diesel', 'Unleaded']:
        print(f"\n  Training LightGBM for {fuel}...")
        model = _train_lgbm(monthly_veh, fuel)

        if model is None:
            print(f"  No training data for {fuel} — skipping.")
            continue

        fuel_eligible = eligible[eligible['fuel_type'] == fuel]['reg'].unique()
        fuel_data     = monthly_veh[monthly_veh['fuel_type'] == fuel]

        for reg in fuel_eligible:
            veh_data = (
                fuel_data[fuel_data['reg'] == reg]
                .sort_values('year_month')
                .set_index('year_month')
            )
            history = veh_data['litres'].astype(float)

            branch     = veh_data['branch'].iloc[-1]
            asset_type = veh_data['asset_type'].iloc[-1] if 'asset_type' in veh_data.columns else ''
            make       = veh_data['make'].iloc[-1]        if 'make' in veh_data.columns else ''
            model_name = veh_data['model'].iloc[-1]       if 'model' in veh_data.columns else ''

            preds = _forecast_vehicle(history, model, steps=12)

            for i, (month, pred_litres) in enumerate(zip(FORECAST_MONTHS, preds)):
                all_forecasts.append({
                    'year_month':  str(month),
                    'reg':         reg,
                    'branch':      branch,
                    'fuel_type':   fuel,
                    'asset_type':  asset_type,
                    'make':        make,
                    'model':       model_name,
                    'pred_litres': round(pred_litres, 2),
                })

        print(f"  {fuel}: forecasted {len(fuel_eligible)} vehicles × 12 months")

    forecasts_df = pd.DataFrame(all_forecasts)

    # Apply EV fleet adjustment: compound monthly volume reduction
    # step_index is 1-based (month 1 = first forecast month)
    forecasts_df['step_index'] = forecasts_df['year_month'].apply(
        lambda m: list([str(p) for p in FORECAST_MONTHS]).index(m) + 1
    )
    monthly_ev_factor = (1 - EV_ANNUAL_VOLUME_REDUCTION) ** (1 / 12)
    forecasts_df['ev_factor']    = monthly_ev_factor ** forecasts_df['step_index']
    forecasts_df['pred_litres']  = (forecasts_df['pred_litres'] * forecasts_df['ev_factor']).round(2)
    forecasts_df = forecasts_df.drop(columns=['step_index', 'ev_factor'])

    forecasts_df.to_csv(PROCESSED_DIR / 'volume_forecasts.csv', index=False)

    ev_pct = (1 - monthly_ev_factor ** 12) * 100
    print(f"\n  EV fleet adjustment applied: {EV_ANNUAL_VOLUME_REDUCTION:.0%}/year ({ev_pct:.1f}% total over 12 months)")
    print(f"  Total forecast rows: {len(forecasts_df):,}")
    print("Section 3 complete.\n")
    return forecasts_df


# ─────────────────────────────────────────────────────────────────────────────
# BIAS CORRECTION — rolling actual vs predicted
# ─────────────────────────────────────────────────────────────────────────────

BIAS_CORRECTION_MONTHS = 3    # how many verified months to average for the correction

HISTORY_PATH = PROCESSED_DIR / 'forecast_history.csv'


def load_forecast_history() -> pd.DataFrame:
    if HISTORY_PATH.exists():
        return pd.read_csv(HISTORY_PATH)
    return pd.DataFrame(columns=['run_date', 'year_month', 'pred_litres_total'])


def save_forecast_history(forecasts_df: pd.DataFrame):
    """Append this run's company-level totals to the forecast history."""
    history = load_forecast_history()
    run_date = datetime.today().strftime('%Y-%m-%d')

    company_totals = (
        forecasts_df.groupby('year_month')['pred_litres']
        .sum()
        .reset_index()
        .rename(columns={'pred_litres': 'pred_litres_total'})
    )
    company_totals['run_date'] = run_date

    # Remove any same-day entries (re-run today) to avoid duplicates
    history = history[history['run_date'] != run_date]
    history = pd.concat([history, company_totals], ignore_index=True)
    history.to_csv(HISTORY_PATH, index=False)


def apply_bias_correction(forecasts_df: pd.DataFrame,
                           monthly_company: pd.DataFrame) -> tuple:
    """
    Compare previously forecast months against actual litres now available.
    Returns (corrected forecasts_df, bias_df) where bias_df is the comparison
    table for the report.  correction_factor of 1.0 = no adjustment.
    """
    history = load_forecast_history()

    # Actual company-level litres per month (all fuel types combined)
    actual = monthly_company.copy()
    actual['year_month'] = actual['year_month'].apply(
        lambda x: str(pd.Period(str(x), freq='M'))
    )
    actual_totals = actual.groupby('year_month')['litres'].sum().reset_index()
    actual_totals.columns = ['year_month', 'actual_litres']

    bias_df = pd.DataFrame()
    correction_factor = 1.0

    if len(history) == 0:
        print("  No forecast history yet — bias correction will activate after first forecast month is verified.")
        return forecasts_df, bias_df, correction_factor

    # For each month that has been forecast, use the earliest forecast
    # (the first time we predicted it, before actuals arrived)
    hist_first = (
        history.sort_values('run_date')
        .groupby('year_month')[['run_date', 'pred_litres_total']]
        .first()
        .reset_index()
    )

    # Inner join: only months where we have both a prior forecast and actual data
    comparison = hist_first.merge(actual_totals, on='year_month', how='inner')

    if len(comparison) == 0:
        print("  No verified forecast months yet — bias correction not yet active.")
        return forecasts_df, bias_df, correction_factor

    comparison['bias_ratio']   = comparison['actual_litres'] / comparison['pred_litres_total']
    comparison['variance_pct'] = (comparison['bias_ratio'] - 1) * 100
    comparison = comparison.sort_values('year_month')

    # Rolling average of last BIAS_CORRECTION_MONTHS verified months
    recent = comparison.tail(BIAS_CORRECTION_MONTHS)
    correction_factor = float(recent['bias_ratio'].mean())

    direction = "lower" if correction_factor < 1.0 else "higher"
    print(f"\n  Bias correction ({len(recent)} verified months):")
    for _, row in recent.iterrows():
        print(f"    {row['year_month']}: actual={row['actual_litres']:,.0f}L  "
              f"predicted={row['pred_litres_total']:,.0f}L  "
              f"variance={row['variance_pct']:+.1f}%")
    print(f"  Correction factor: {correction_factor:.4f} "
          f"(actuals running {abs(correction_factor - 1):.1%} {direction} than forecast)")

    # Sanity check: ignore extreme corrections (data issue rather than model drift)
    if not (0.50 <= correction_factor <= 1.50):
        print(f"  WARNING: correction factor {correction_factor:.4f} outside [0.5, 1.5] — not applied.")
        correction_factor = 1.0
    else:
        forecasts_df = forecasts_df.copy()
        forecasts_df['pred_litres'] = (forecasts_df['pred_litres'] * correction_factor).round(2)

    bias_df = comparison[['year_month', 'run_date', 'pred_litres_total',
                           'actual_litres', 'bias_ratio', 'variance_pct']].copy()
    return forecasts_df, bias_df, correction_factor


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 4 — REPORT OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

# ── Style helpers ─────────────────────────────────────────────────────────────

DARK_BLUE  = 'FF1F3864'
MID_BLUE   = 'FF2E75B6'
LIGHT_BLUE = 'FFD6E4F0'
PALE_GREY  = 'FFF2F2F2'
WHITE      = 'FFFFFFFF'
AMBER      = 'FFFFC000'
RED_LIGHT  = 'FFFF0000'

def _hdr_font(size=11, bold=True, color='FFFFFFFF'):
    return Font(name='Calibri', size=size, bold=bold, color=color)

def _body_font(size=10, bold=False, color='FF000000'):
    return Font(name='Calibri', size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _thin_border():
    s = Side(style='thin', color='FFB0B0B0')
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def _right():
    return Alignment(horizontal='right', vertical='center')

def _fmt_gbp(val):
    if pd.isna(val):
        return ''
    return f'£{val:,.0f}'

def _fmt_pct(val):
    if pd.isna(val):
        return ''
    return f'{val:.1%}'

def _fmt_litres(val):
    if pd.isna(val):
        return ''
    return f'{val:,.0f}'

def _apply_header_row(ws, row_idx, values, col_start=1,
                      bg=DARK_BLUE, fg='FFFFFFFF', height=22):
    ws.row_dimensions[row_idx].height = height
    for i, val in enumerate(values):
        c = ws.cell(row=row_idx, column=col_start + i, value=val)
        c.font      = _hdr_font(color=fg)
        c.fill      = _fill(bg)
        c.alignment = _center()
        c.border    = _thin_border()


def _set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_company_sheet(ws, company_hist: pd.DataFrame,
                          company_fcst: pd.DataFrame,
                          price_proj: pd.DataFrame):
    """Company-level overview: actual history + 12-month forecast."""
    ws.title = 'Company Overview'

    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = 'Company Fuel Forecast — April 2026 to March 2027'
    ws['A1'].font      = _hdr_font(size=14)
    ws['A1'].fill      = _fill(DARK_BLUE)
    ws['A1'].alignment = _center()
    ws.row_dimensions[1].height = 28

    # Sub-header
    ws.merge_cells('A2:G2')
    ws['A2'] = f'Generated {datetime.today().strftime("%d %B %Y")}'
    ws['A2'].font      = _body_font(size=9, color='FF666666')
    ws['A2'].alignment = _center()

    # ── Section 1: 12-month forecast ─────────────────────────────
    ws.merge_cells('A4:G4')
    ws['A4'] = 'Forecast — April 2026 to March 2027'
    ws['A4'].font      = _hdr_font(size=11, color='FF1F3864')
    ws['A4'].fill      = _fill(LIGHT_BLUE)
    ws['A4'].alignment = Alignment(horizontal='left', vertical='center')

    fcst_headers = ['Month', 'Fuel Type', 'Predicted Litres', 'Price £/Litre', 'Predicted Spend (£)']
    _apply_header_row(ws, 5, fcst_headers, bg=MID_BLUE)

    # Price lookup
    pp = price_proj.set_index('year_month')

    r = 6
    for _, row in company_fcst.sort_values(['year_month', 'fuel_type']).iterrows():
        price = (pp.loc[row['year_month'], 'diesel_price']
                 if row['fuel_type'] == 'Diesel'
                 else pp.loc[row['year_month'], 'unleaded_price'])
        spend = row['pred_litres'] * price
        bg    = PALE_GREY if r % 2 == 0 else WHITE
        vals  = [row['year_month'], row['fuel_type'],
                 _fmt_litres(row['pred_litres']),
                 f"£{price:.4f}",
                 _fmt_gbp(spend)]
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=1 + i, value=v)
            c.font      = _body_font()
            c.fill      = _fill(bg)
            c.alignment = _center() if i > 1 else Alignment(horizontal='left', vertical='center')
            c.border    = _thin_border()
        r += 1

    # ── Section 2: Historical actuals (last 6 months) ────────────
    r += 1
    ws.merge_cells(f'A{r}:G{r}')
    ws[f'A{r}'] = 'Historical Actuals (last 6 months)'
    ws[f'A{r}'].font      = _hdr_font(size=11, color='FF1F3864')
    ws[f'A{r}'].fill      = _fill(LIGHT_BLUE)
    ws[f'A{r}'].alignment = Alignment(horizontal='left', vertical='center')
    r += 1

    hist_headers = ['Month', 'Fuel Type', 'Litres', 'Net Amount (£)', 'Avg £/Litre']
    _apply_header_row(ws, r, hist_headers, bg=MID_BLUE)
    r += 1

    hist6 = company_hist.copy()
    hist6['ym'] = hist6['year_month'].apply(lambda x: pd.Period(str(x), freq='M'))
    hist6 = hist6.sort_values('ym')
    hist6 = hist6.groupby('fuel_type', group_keys=False).tail(6)
    hist6 = hist6.sort_values(['ym', 'fuel_type'])

    for _, row in hist6.iterrows():
        bg   = PALE_GREY if r % 2 == 0 else WHITE
        vals = [str(row['ym']), row['fuel_type'],
                _fmt_litres(row['litres']),
                _fmt_gbp(row['net_amount']),
                f"£{row['avg_unit_price']:.4f}" if not pd.isna(row['avg_unit_price']) else '']
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=1 + i, value=v)
            c.font      = _body_font()
            c.fill      = _fill(bg)
            c.alignment = _center() if i > 1 else Alignment(horizontal='left', vertical='center')
            c.border    = _thin_border()
        r += 1

    _set_col_widths(ws, {'A': 16, 'B': 14, 'C': 18, 'D': 14, 'E': 20, 'F': 14, 'G': 14})
    ws.freeze_panes = 'A6'


def _build_branch_sheet(ws, branch_hist: pd.DataFrame,
                         branch_fcst: pd.DataFrame,
                         price_proj: pd.DataFrame):
    """Branch-level forecast summary: one row per branch×fuel per forecast month."""
    ws.title = 'Branch Forecast'

    pp = price_proj.set_index('year_month')
    months = [str(m) for m in FORECAST_MONTHS]

    # Build summary: total predicted spend per branch×fuel×month
    branch_fcst = branch_fcst.copy()
    branch_fcst['price'] = branch_fcst.apply(
        lambda r: pp.loc[r['year_month'], 'diesel_price']
        if r['fuel_type'] == 'Diesel'
        else pp.loc[r['year_month'], 'unleaded_price'],
        axis=1
    )
    branch_fcst['pred_spend'] = branch_fcst['pred_litres'] * branch_fcst['price']

    # Combine diesel + unleaded: one row per branch
    pivot_spend = (
        branch_fcst.groupby(['branch', 'year_month'])['pred_spend']
        .sum()
        .unstack('year_month')
        .reindex(columns=months, fill_value=0)
    )
    pivot_spend['Total Spend (£)'] = pivot_spend.sum(axis=1)
    pivot_spend = pivot_spend.sort_values('Total Spend (£)', ascending=False)

    n_months = len(months)
    n_cols   = 1 + n_months + 1   # Branch, 12 months, Total

    # Title
    ws.merge_cells(f'A1:{get_column_letter(n_cols)}1')
    ws['A1'] = 'Branch Fuel Forecast — April 2026 to March 2027'
    ws['A1'].font      = _hdr_font(size=14)
    ws['A1'].fill      = _fill(DARK_BLUE)
    ws['A1'].alignment = _center()
    ws.row_dimensions[1].height = 28

    headers = ['Branch'] + months + ['Total Spend (£)']
    _apply_header_row(ws, 2, headers, bg=MID_BLUE)

    r = 3
    for branch, spend_row in pivot_spend.iterrows():
        bg       = PALE_GREY if r % 2 == 0 else WHITE
        row_vals = [branch]
        for m in months:
            row_vals.append(_fmt_gbp(spend_row.get(m, 0)))
        row_vals.append(_fmt_gbp(spend_row.get('Total Spend (£)', 0)))

        for i, v in enumerate(row_vals):
            c = ws.cell(row=r, column=1 + i, value=v)
            c.font      = _body_font()
            c.fill      = _fill(bg)
            c.alignment = _right() if i > 0 else Alignment(horizontal='left', vertical='center')
            c.border    = _thin_border()
        r += 1

    ws.freeze_panes = 'B3'
    ws.auto_filter.ref = f'A2:{get_column_letter(n_cols)}2'
    ws.column_dimensions['A'].width = 22
    for i in range(n_months + 1):
        ws.column_dimensions[get_column_letter(2 + i)].width = 14


def _build_vehicle_sheet(ws, veh_fcst: pd.DataFrame, price_proj: pd.DataFrame):
    """Vehicle-level forecast: one row per vehicle×fuel×month."""
    ws.title = 'Vehicle Forecast'

    pp = price_proj.set_index('year_month')
    months = [str(m) for m in FORECAST_MONTHS]

    # Compute spend
    veh_fcst = veh_fcst.copy()
    veh_fcst['price'] = veh_fcst.apply(
        lambda r: pp.loc[r['year_month'], 'diesel_price']
        if r['fuel_type'] == 'Diesel'
        else pp.loc[r['year_month'], 'unleaded_price'],
        axis=1
    )
    veh_fcst['pred_spend'] = veh_fcst['pred_litres'] * veh_fcst['price']

    # Combine diesel + unleaded: one row per vehicle
    # Use last-known branch/make/model for each reg
    veh_meta = (
        veh_fcst.sort_values('year_month')
        .groupby('reg')[['branch', 'make', 'model']]
        .last()
        .reset_index()
    )
    pivot = (
        veh_fcst.groupby(['reg', 'year_month'])['pred_spend']
        .sum()
        .unstack('year_month')
        .reindex(columns=months, fill_value=0)
        .reset_index()
    )
    pivot = pivot.merge(veh_meta, on='reg', how='left')
    pivot['Total Spend (£)'] = pivot[months].sum(axis=1)
    pivot = pivot.sort_values(['branch', 'Total Spend (£)'], ascending=[True, False])

    # Title
    n_cols = 4 + len(months) + 1   # Reg, Branch, Make, Model, months, Total
    ws.merge_cells(f'A1:{get_column_letter(n_cols)}1')
    ws['A1'] = 'Vehicle Fuel Forecast — April 2026 to March 2027'
    ws['A1'].font      = _hdr_font(size=14)
    ws['A1'].fill      = _fill(DARK_BLUE)
    ws['A1'].alignment = _center()
    ws.row_dimensions[1].height = 28

    headers = ['Registration', 'Branch', 'Make', 'Model'] + months + ['Total Spend (£)']
    _apply_header_row(ws, 2, headers, bg=MID_BLUE)

    r = 3
    for _, row in pivot.iterrows():
        bg = PALE_GREY if r % 2 == 0 else WHITE
        vals = [row['reg'], row.get('branch', ''),
                row.get('make', ''), row.get('model', '')]
        for m in months:
            vals.append(_fmt_gbp(row.get(m, 0)))
        vals.append(_fmt_gbp(row.get('Total Spend (£)', 0)))

        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=1 + i, value=v)
            c.font      = _body_font()
            c.fill      = _fill(bg)
            c.alignment = _right() if i >= 4 else Alignment(horizontal='left', vertical='center')
            c.border    = _thin_border()
        r += 1

    ws.freeze_panes = 'E3'
    ws.auto_filter.ref = f'A2:{get_column_letter(n_cols)}2'

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    for i in range(len(months) + 1):
        ws.column_dimensions[get_column_letter(5 + i)].width = 14


def _build_assumptions_sheet(ws, price_proj: pd.DataFrame,
                               diesel_chg: float, unl_chg: float,
                               bias_df: pd.DataFrame = None,
                               correction_factor: float = 1.0):
    """Assumptions & methodology sheet."""
    ws.title = 'Assumptions'

    ws.merge_cells('A1:E1')
    ws['A1'] = 'Forecast Assumptions & Methodology'
    ws['A1'].font      = _hdr_font(size=14)
    ws['A1'].fill      = _fill(DARK_BLUE)
    ws['A1'].alignment = _center()
    ws.row_dimensions[1].height = 28

    notes = [
        ('Volume Model',
         'LightGBM gradient boosting — panel model trained across all vehicles with ≥3 months history. '
         'Features: lags 1/2/3/6 months, rolling averages 3/6 months, month seasonality (sin/cos). '
         'Recursive 12-step forecast per vehicle.'),
        ('EV Fleet Adjustment',
         f'A {EV_ANNUAL_VOLUME_REDUCTION:.0%}/year volume reduction is applied on top of the LightGBM predictions '
         f'to reflect the ongoing electrification of the fleet (company target: 15%/year). '
         f'Applied as a compounding monthly factor ({(1-(1-EV_ANNUAL_VOLUME_REDUCTION)**(1/12)):.2%}/month), '
         f'reducing forecast volumes by ~{EV_ANNUAL_VOLUME_REDUCTION:.0%} over the 12-month period. '
         f'Adjust EV_ANNUAL_VOLUME_REDUCTION in run_pipeline.py as fleet electrification data becomes clearer.'),
        ('Price Model',
         f'March 2026 Shell unit price used as anchor (Diesel £{PRICE_ANCHOR["Diesel"]:.4f}/L, '
         f'Unleaded £{PRICE_ANCHOR["Unleaded"]:.4f}/L). '
         f'BEIS/DESNZ weekly pump price trend applied: '
         f'Diesel {diesel_chg:+.2%}/month, Unleaded {unl_chg:+.2%}/month (3-month average).'),
        ('Card-only transactions',
         'Transactions without a vehicle registration are included in Branch and Company totals '
         'but excluded from vehicle-level forecasts.'),
        ('Fuel types',
         'Diesel + V Power Diesel grouped as Diesel. '
         'Unleaded Medium Octane + Unleaded High Perf grouped as Unleaded.'),
        ('Forecast period',
         'April 2026 – March 2027 (12 months).'),
    ]

    r = 3
    for title, body in notes:
        ws.cell(row=r, column=1).value = title
        ws.cell(row=r, column=1).font  = _hdr_font(size=11, color='FF1F3864')
        ws.cell(row=r, column=1).fill  = _fill(LIGHT_BLUE)
        ws.merge_cells(f'A{r}:E{r}')
        r += 1

        ws.cell(row=r, column=1).value     = body
        ws.cell(row=r, column=1).font      = _body_font()
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[r].height        = 45
        ws.merge_cells(f'A{r}:E{r}')
        r += 2

    # Price projection table
    r += 1
    ws.cell(row=r, column=1).value = 'Projected Unit Prices (£/Litre)'
    ws.cell(row=r, column=1).font  = _hdr_font(size=11, color='FF1F3864')
    ws.cell(row=r, column=1).fill  = _fill(LIGHT_BLUE)
    ws.merge_cells(f'A{r}:C{r}')
    r += 1
    _apply_header_row(ws, r, ['Month', 'Diesel (£/L)', 'Unleaded (£/L)'], bg=MID_BLUE)
    r += 1

    for _, prow in price_proj.iterrows():
        ws.cell(row=r, column=1).value = prow['year_month']
        ws.cell(row=r, column=2).value = prow['diesel_price']
        ws.cell(row=r, column=3).value = prow['unleaded_price']
        for col in [1, 2, 3]:
            ws.cell(row=r, column=col).font   = _body_font()
            ws.cell(row=r, column=col).border = _thin_border()
        r += 1

    # Bias correction table
    r += 1
    ws.merge_cells(f'A{r}:F{r}')
    ws[f'A{r}'] = 'Forecast Accuracy — Actual vs Predicted (company volume)'
    ws[f'A{r}'].font      = _hdr_font(size=11, color='FF1F3864')
    ws[f'A{r}'].fill      = _fill(LIGHT_BLUE)
    ws[f'A{r}'].alignment = Alignment(horizontal='left', vertical='center')
    r += 1

    if bias_df is not None and len(bias_df) > 0:
        _apply_header_row(ws, r,
            ['Month', 'First Forecast Date', 'Predicted Litres',
             'Actual Litres', 'Variance %', 'Bias Ratio'],
            bg=MID_BLUE)
        r += 1
        for _, brow in bias_df.iterrows():
            variance = brow['variance_pct']
            bg = PALE_GREY if r % 2 == 0 else WHITE
            # Colour the variance cell: green if within 5%, amber 5-15%, red >15%
            if abs(variance) <= 5:
                var_bg = 'FF92D050'   # green
            elif abs(variance) <= 15:
                var_bg = 'FFFFC000'   # amber
            else:
                var_bg = 'FFFF0000'   # red

            row_vals = [brow['year_month'], brow['run_date'],
                        _fmt_litres(brow['pred_litres_total']),
                        _fmt_litres(brow['actual_litres']),
                        f"{variance:+.1f}%",
                        f"{brow['bias_ratio']:.4f}"]
            for i, v in enumerate(row_vals):
                c = ws.cell(row=r, column=1 + i, value=v)
                c.font      = _body_font()
                c.fill      = _fill(var_bg if i == 4 else bg)
                c.alignment = _center()
                c.border    = _thin_border()
            r += 1

        # Summary row
        r += 1
        ws.cell(row=r, column=1).value = f'Applied correction factor (last {min(BIAS_CORRECTION_MONTHS, len(bias_df))} months):'
        ws.cell(row=r, column=1).font  = _hdr_font(size=10, color='FF1F3864')
        ws.merge_cells(f'A{r}:D{r}')
        ws.cell(row=r, column=5).value = f'{(correction_factor - 1) * 100:+.2f}%'
        ws.cell(row=r, column=6).value = f'{correction_factor:.4f}'
        for col in [5, 6]:
            ws.cell(row=r, column=col).font      = _hdr_font(size=10)
            ws.cell(row=r, column=col).alignment = _center()
    else:
        ws.cell(row=r, column=1).value = (
            'No verified forecast months yet. '
            'Once a previously forecast month has actual data loaded, '
            'the variance will appear here and auto-correct future forecasts.'
        )
        ws.cell(row=r, column=1).font      = _body_font(color='FF666666')
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[r].height        = 35
        ws.merge_cells(f'A{r}:F{r}')

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14


def section4_report(monthly_veh: pd.DataFrame,
                     monthly_branch: pd.DataFrame,
                     monthly_company: pd.DataFrame,
                     forecasts_df: pd.DataFrame,
                     price_proj: pd.DataFrame,
                     bias_df: pd.DataFrame = None,
                     correction_factor: float = 1.0):
    print("=" * 60)
    print("SECTION 4 — Report Output")
    print("=" * 60)

    # ── Aggregate forecasts to branch / company ───────────────────
    branch_fcst = (
        forecasts_df.groupby(['year_month', 'branch', 'fuel_type'])
        ['pred_litres']
        .sum()
        .reset_index()
    )

    company_fcst = (
        forecasts_df.groupby(['year_month', 'fuel_type'])
        ['pred_litres']
        .sum()
        .reset_index()
    )

    # Price change values used (read back from CSV for assumptions sheet)
    pp = price_proj.copy()
    diesel_changes   = []
    unleaded_changes = []
    for i in range(1, len(pp)):
        diesel_changes.append(pp['diesel_price'].iloc[i] / pp['diesel_price'].iloc[i - 1] - 1)
        unleaded_changes.append(pp['unleaded_price'].iloc[i] / pp['unleaded_price'].iloc[i - 1] - 1)
    diesel_chg   = float(np.mean(diesel_changes))   if diesel_changes   else 0.0025
    unleaded_chg = float(np.mean(unleaded_changes)) if unleaded_changes else 0.0020

    # ── Build workbook ────────────────────────────────────────────
    wb = Workbook()
    ws_company = wb.active
    ws_branch  = wb.create_sheet()
    ws_vehicle = wb.create_sheet()
    ws_assump  = wb.create_sheet()

    _build_company_sheet(ws_company, monthly_company, company_fcst, price_proj)
    _build_branch_sheet(ws_branch,  monthly_branch,  branch_fcst,  price_proj)
    _build_vehicle_sheet(ws_vehicle, forecasts_df,   price_proj)
    _build_assumptions_sheet(ws_assump, price_proj, diesel_chg, unleaded_chg,
                              bias_df=bias_df, correction_factor=correction_factor)

    # ── Save ──────────────────────────────────────────────────────
    today    = datetime.today()
    filename = f'Fuel_Forecast_{today.strftime("%Y%m%d_%H%M%S")}.xlsx'
    out_path = REPORTS_DIR / filename
    wb.save(out_path)

    print(f"\n  Report saved: {out_path}")
    print("Section 4 complete.\n")
    return out_path


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    start = datetime.now()

    monthly_veh, monthly_branch, monthly_company = section1_load_data()
    price_proj   = section2_price_projection()
    forecasts_df = section3_volume_prediction(monthly_veh)

    # Bias correction: compare previous forecasts against actuals now available
    print("=" * 60)
    print("BIAS CORRECTION — Actual vs Predicted")
    print("=" * 60)
    forecasts_df, bias_df, correction_factor = apply_bias_correction(
        forecasts_df, monthly_company
    )

    # Save this run's forecasts for future verification
    save_forecast_history(forecasts_df)
    print(f"  Forecast history updated: {HISTORY_PATH}")

    out_path = section4_report(
        monthly_veh, monthly_branch, monthly_company,
        forecasts_df, price_proj,
        bias_df=bias_df, correction_factor=correction_factor
    )

    elapsed = (datetime.now() - start).total_seconds()
    print(f"\nPipeline complete in {elapsed:.1f}s")
    print(f"Report: {out_path}")
